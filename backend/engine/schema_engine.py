"""
engine/schema_engine.py
-----------------------
Dynamically creates new modules from plain English descriptions.
When you say "I want to track my freelance clients", this file:
  1. Asks Llama to design the schema
  2. Saves it to modules.json
  3. Creates the SQLite table
  4. Creates the Excel file
"""

import json
from pathlib import Path

from backend.core import brain, safeguards


# ─────────────────────────────────────────────
#  Paths
# ─────────────────────────────────────────────

ROOT         = Path(__file__).parent.parent.parent
CONFIG_PATH  = ROOT / "config" / "modules.json"
PROMPT_PATH  = ROOT / "config" / "prompts" / "schema_prompt.txt"


# ─────────────────────────────────────────────
#  Load / Save modules.json
# ─────────────────────────────────────────────

def load_config() -> dict:
    with open(CONFIG_PATH, "r") as f:
        return json.load(f)


def save_config(config: dict):
    with open(CONFIG_PATH, "w") as f:
        json.dump(config, f, indent=2)


def get_modules() -> dict:
    return load_config().get("modules", {})


def get_module(name: str) -> dict | None:
    return get_modules().get(name)


def module_exists(name: str) -> bool:
    return name in get_modules()


# ─────────────────────────────────────────────
#  Create a new module from plain English
# ─────────────────────────────────────────────

def create_module(description: str, module_key: str = None) -> dict:
    """
    Create a brand new module from a plain English description.

    Args:
        description: e.g. "Track my freelance clients and their projects"
        module_key:  Optional. If None, AI will suggest a key name.

    Returns:
        The created module schema dict.
    """

    # Step 1 — Ask Llama to design the schema
    schema = _generate_schema(description)

    # Step 2 — Determine the module key (short snake_case name)
    if not module_key:
        module_key = _generate_key(description, schema.get("label", ""))

    # Step 3 — Check for conflicts
    if module_exists(module_key):
        raise ValueError(
            f"Module '{module_key}' already exists. "
            f"Choose a different name or update the existing one."
        )

    # Step 4 — Add metadata
    schema["excel_file"] = f"{module_key}.xlsx"
    schema["sources"]    = schema.get("sources", ["manual"])
    schema["auto_scan"]  = False
    schema["schedule"]   = None

    # Step 5 — Save to modules.json
    config = load_config()
    config["modules"][module_key] = schema
    save_config(config)

    # Step 6 — Create the Excel file
    _create_excel_file(module_key, schema)

    print(f"Module '{module_key}' created successfully.")
    print(f"  Label  : {schema.get('label')}")
    print(f"  Fields : {[f['name'] for f in schema.get('fields', [])]}")
    print(f"  Excel  : {schema['excel_file']}")

    return {module_key: schema}


# ─────────────────────────────────────────────
#  Update an existing module
# ─────────────────────────────────────────────

def add_field(module_key: str, field: dict):
    """
    Add a new field to an existing module.

    Args:
        module_key: The module name e.g. "jobs"
        field: {"name": "priority", "type": "enum",
                "required": false, "options": ["low","high"]}
    """
    if not module_exists(module_key):
        raise ValueError(f"Module '{module_key}' does not exist.")

    config = load_config()
    fields = config["modules"][module_key]["fields"]

    # Check field doesn't already exist
    existing_names = [f["name"] for f in fields]
    if field["name"] in existing_names:
        raise ValueError(
            f"Field '{field['name']}' already exists in module '{module_key}'."
        )

    fields.append(field)
    config["modules"][module_key]["fields"] = fields
    save_config(config)

    # Update Excel file with new column
    _add_excel_column(module_key, field["name"])

    print(f"Field '{field['name']}' added to module '{module_key}'.")


def remove_field(module_key: str, field_name: str):
    """Remove a field from an existing module."""
    if not module_exists(module_key):
        raise ValueError(f"Module '{module_key}' does not exist.")

    config = load_config()
    fields = config["modules"][module_key]["fields"]
    config["modules"][module_key]["fields"] = [
        f for f in fields if f["name"] != field_name
    ]
    save_config(config)
    print(f"Field '{field_name}' removed from module '{module_key}'.")


def update_module(module_key: str, schema: dict) -> dict:
    """Update a module definition while preserving its key and existing data."""
    if not module_exists(module_key):
        raise ValueError(f"Module '{module_key}' does not exist.")
    schema = _validate_schema(schema)
    config = load_config()
    previous = config["modules"][module_key]
    schema["excel_file"] = previous.get("excel_file", f"{module_key}.xlsx")
    schema["sources"] = schema.get("sources", previous.get("sources", ["manual"]))
    schema["auto_scan"] = schema.get("auto_scan", previous.get("auto_scan", False))
    schema["schedule"] = schema.get("schedule", previous.get("schedule"))
    config["modules"][module_key] = schema
    save_config(config)
    return schema


def delete_module(module_key: str):
    """Delete a module entirely."""
    if not module_exists(module_key):
        raise ValueError(f"Module '{module_key}' does not exist.")

    config = load_config()
    del config["modules"][module_key]
    save_config(config)
    print(f"Module '{module_key}' deleted.")


def create_module_from_schema(module_key: str, schema: dict) -> dict:
    """Create a module from an explicit schema without asking the model."""
    if module_exists(module_key):
        raise ValueError(f"Module '{module_key}' already exists.")
    schema = _validate_schema(schema)
    schema["excel_file"] = schema.get("excel_file", f"{module_key}.xlsx")
    schema["sources"] = schema.get("sources", ["manual"])
    schema["auto_scan"] = schema.get("auto_scan", False)
    schema["schedule"] = schema.get("schedule")
    config = load_config()
    config["modules"][module_key] = schema
    save_config(config)
    _create_excel_file(module_key, schema)
    return {module_key: schema}


# ─────────────────────────────────────────────
#  AI schema generation
# ─────────────────────────────────────────────

def _generate_schema(description: str) -> dict:
    """Ask Llama to design a schema for the described module."""

    prompt = (
        "Design a data schema for this use case. Return JSON only.\n\n"
        "Use case:\n" + safeguards.wrap_user_text(description, "MODULE_DESCRIPTION") + "\n\n"
        "Return this exact JSON structure:\n"
        '{"label": "Human Name", "icon": "emoji", '
        '"description": "what it tracks", '
        '"fields": [{"name": "field_name", "type": "text|number|date|enum|boolean", '
        '"required": true, "options": ["only if enum"]}], '
        '"sources": ["manual"]}\n\n'
        "Rules:\n"
        "- field names must be lowercase with underscores\n"
        "- always include a date field\n"
        "- always include a status or notes field\n"
        "- for enum fields always include options array\n"
        "- 5 to 10 fields is ideal\n"
        "- return raw JSON only, no explanation"
    )

    schema = brain.ask_json(prompt, temperature=0.2)
    return _validate_schema(schema)


def _validate_schema(schema: dict) -> dict:
    """Validate the minimum contract needed before creating tables/files."""
    if not isinstance(schema, dict):
        raise ValueError("Generated schema must be a JSON object.")

    required_top_level = ["label", "description", "fields"]
    missing = [key for key in required_top_level if key not in schema]
    if missing:
        raise ValueError("Generated schema missing required keys: " + ", ".join(missing))

    fields = schema.get("fields")
    if not isinstance(fields, list) or not fields:
        raise ValueError("Generated schema must include at least one field.")

    valid_types = {"text", "number", "date", "enum", "boolean"}
    for field in fields:
        if not isinstance(field, dict):
            raise ValueError("Each generated field must be a JSON object.")

        for key in ["name", "type", "required"]:
            if key not in field:
                raise ValueError("Generated field missing required key: " + key)

        name = field["name"]
        if not isinstance(name, str) or not name or not name.replace("_", "").isalnum():
            raise ValueError("Generated field has invalid name: " + str(name))
        if name.lower() != name or " " in name:
            raise ValueError("Generated field names must be lowercase snake_case: " + name)

        field_type = field["type"]
        if field_type not in valid_types:
            raise ValueError("Generated field has invalid type: " + str(field_type))
        if field_type == "enum" and not field.get("options"):
            raise ValueError("Enum field must include options: " + name)
        if not isinstance(field["required"], bool):
            raise ValueError("Generated field 'required' must be boolean: " + name)

    return schema


def _generate_key(description: str, label: str) -> str:
    """Generate a short snake_case key for the module."""

    prompt = (
        "Generate a short snake_case identifier for this module.\n"
        "Description:\n" + safeguards.wrap_user_text(description, "MODULE_DESCRIPTION") + "\n"
        "Label: " + label + "\n\n"
        "Rules:\n"
        "- lowercase only\n"
        "- underscores instead of spaces\n"
        "- 1 to 3 words maximum\n"
        "- no special characters\n\n"
        'Return only the key string, nothing else. Example: "freelance_clients"'
    )

    try:
        key = brain.ask(prompt, temperature=0.1).strip().lower()
        # Clean up — remove quotes, spaces, special chars
        key = key.replace('"', '').replace("'", '').replace(' ', '_')
        key = ''.join(c for c in key if c.isalnum() or c == '_')
        return key or "new_module"
    except Exception:
        # Fallback — generate from label
        return label.lower().replace(' ', '_')[:20] or "new_module"


# ─────────────────────────────────────────────
#  Excel file creation
# ─────────────────────────────────────────────

def _create_excel_file(module_key: str, schema: dict):
    """Create an Excel file with headers for this module."""
    try:
        import openpyxl
        excel_dir = ROOT / "data" / "excel"
        excel_dir.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = schema.get("label", module_key)

        # Write headers from field names
        headers = ["id"] + [f["name"] for f in schema.get("fields", [])] + ["created_at"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)

        filepath = excel_dir / schema.get("excel_file", f"{module_key}.xlsx")
        wb.save(filepath)
        print(f"  Excel  : {filepath}")

    except ImportError:
        print("  Excel  : openpyxl not installed, skipping Excel creation")


def _add_excel_column(module_key: str, field_name: str):
    """Add a new column to an existing Excel file."""
    try:
        import openpyxl
        schema    = get_module(module_key)
        excel_dir = ROOT / "data" / "excel"
        filepath  = excel_dir / schema.get("excel_file", f"{module_key}.xlsx")

        if not filepath.exists():
            _create_excel_file(module_key, schema)
            return

        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        next_col = ws.max_column + 1
        cell = ws.cell(row=1, column=next_col, value=field_name)
        cell.font = openpyxl.styles.Font(bold=True)
        wb.save(filepath)

    except ImportError:
        print("openpyxl not installed, skipping Excel update")


# ─────────────────────────────────────────────
#  Quick test
# ─────────────────────────────────────────────

if __name__ == "__main__":
    print("Testing schema engine...\n")

    # Test 1 — Generate a schema without saving
    print("Test 1: Generate schema for 'freelance clients'")
    print("-" * 40)
    schema = _generate_schema("Track my freelance clients, their projects, rates, and payment status")
    print(f"  Label  : {schema.get('label')}")
    print(f"  Icon   : {schema.get('icon')}")
    print(f"  Fields : {[f['name'] for f in schema.get('fields', [])]}")

    # Test 2 — Generate a module key
    print("\nTest 2: Generate module key")
    print("-" * 40)
    key = _generate_key("Track my freelance clients", schema.get("label", ""))
    print(f"  Key    : {key}")

    # Test 3 — List existing modules
    print("\nTest 3: Existing modules")
    print("-" * 40)
    modules = get_modules()
    for name, mod in modules.items():
        print(f"  {mod.get('icon', '')} {name}: {mod.get('description', '')}")
